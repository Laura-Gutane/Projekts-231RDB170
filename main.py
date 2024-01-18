from openpyxl import Workbook, load_workbook
wb=load_workbook("Izdevumi.xlsx")
ws=wb.active

planots_kopa=0
izterets_kopa=0
starpiba_kopa=0

for row in range(3, 27):
    planots=ws["C"+str(row)].value
    izterets=ws["D"+str(row)].value
    
    if (type(planots)!=str and type(izterets)!=str):
        starpiba=planots-izterets
        ws["E"+str(row)].value=starpiba

        planots_kopa+=planots
        izterets_kopa+=izterets
        starpiba_kopa+=starpiba

        ws["C27"].value=planots_kopa
        ws["D27"].value=izterets_kopa
        ws["E27"].value=starpiba_kopa



wb.save("Izdevumi_rezultats.xlsx")
wb.close